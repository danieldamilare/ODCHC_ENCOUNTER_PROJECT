from app import app
from flask import redirect, flash, url_for, request, render_template, abort, stream_with_context, Response
from flask_login import login_required, login_user, logout_user
from app.models import Role, is_logged_in, get_current_user, AuthUser, Facility, Encounter, User, Disease, TreatmentOutcome, InsuranceScheme, ANCRegistry, Service
from app.services import UserServices, EncounterServices, FacilityServices, DiseaseServices, TreatmentOutcomeServices, ServiceCategoryServices
from app.services import DiseaseCategoryServices, InsuranceSchemeServices, ServiceServices, DownloadServices
from app.exceptions import AuthenticationError, MissingError, ValidationError
from app.exceptions import InvalidReferenceError, DuplicateError, ServiceError
from urllib.parse import urlparse
from app.config import Config
from app.utils import form_to_dict, admin_required, humanize_datetime_filter, calculate_gestational_age, scheme_access_required, get_age_group, autofit_columns, build_filter, parse_date
from app.forms import LoginForm, AddEncounterForm, AddFacilityForm, EditFacilityForm, AddDiseaseForm, ExcelUploadForm, DashboardFilterForm, EncTypeForm, DeliveryEncounterForm, AddServiceForm
from app.forms import AddUserForm, AddCategoryForm, DeleteUserForm, EditUserForm, EditDiseaseForm, EncounterFilterForm, AdminDashboardFilterForm, ANCEncounterForm, ChildHealthEncounterForm, FacilityFilterForm
from app.constants import ONDO_LGAS_LIST, SchemeEnum, BabyOutcome
from werkzeug.utils import secure_filename

from app.filter_parser import Params
from flask_wtf import FlaskForm
from copy import copy
from app.services import DashboardServices, ReportServices, ChatServices
from typing import Optional, List, Dict
from datetime import datetime, date, timedelta
from typing import Any
import json
import io
import arrow
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from flask import g
from flask import send_file
from app.filter_map import filter_config, facility_filter_config, encounter_filter_config, download_encounter_filter_config
import json

def get_facility_user_dashboard():
    facility_id = get_current_user().facility.id
    param_filter = Params().where(Facility, 'id', '=', facility_id)
    start_date = datetime.now().date().replace(day=1)
    end_date=datetime.now().date()
    with_date_param = param_filter.where(Encounter, 'date', '>=', start_date)\
                            .where(Encounter, 'date', '<=', end_date)
    total_encounter = DashboardServices.get_total_encounters(param_filter, start_date=start_date, end_date = end_date)
    encounter_gender = DashboardServices.encounter_gender_distribution(with_date_param)
    total_utilization = DashboardServices.get_total_utilization(param_filter, start_date = start_date, end_date = end_date)
    total_mortality = DashboardServices.get_total_death_outcome(param_filter, start_date = start_date, end_date = end_date)
    encounter_age_group = DashboardServices.encounter_age_group_distribution(with_date_param)
    top_cause_of_mortality = DashboardServices.get_top_cause_of_mortality(with_date_param)
    case_fatality = DashboardServices.case_fatality(with_date_param)
    facility_name = get_current_user().facility.name

    return render_template("facility_dashboard.html",
                           title  = f"{facility_name} - Dashboard",
                           start_date = start_date,
                           end_date = end_date,
                           total_encounter = total_encounter,
                           total_utilization = total_utilization,
                           total_mortality = total_mortality,
                           encounter_gender = encounter_gender,
                           encounter_age_group = encounter_age_group,
                           top_cause_of_mortality = top_cause_of_mortality,
                           facility_name = facility_name,
                           case_fatality = case_fatality
                           )

@app.route('/')
@app.route('/index')
@login_required
def index():

    if get_current_user().role.name != 'admin':
        return get_facility_user_dashboard()
    return redirect(url_for('admin_overview'))

@app.route('/auth/login', methods=['GET', 'POST'])
def login() -> Any:
    if is_logged_in():
        return redirect(url_for('index'))
    form = LoginForm()

    if form.validate_on_submit():
        username = form.username.data
        password = form.password.data
        remember_me = form.remember_me.data
        try:
            user = UserServices.get_verified_user(username, password)
            authuser = AuthUser(user)
            login_user(authuser, remember=remember_me)
            next_page = request.args.get('next')

            if not next_page or urlparse(next_page).netloc != '':
                next_page = url_for('index')
            return redirect(next_page)

        except AuthenticationError as e:
            flash(str(e), 'error')
        # except Exception as e:
            # abort(500)
    return render_template('login.html', title='Sign in', form=form)

@app.route('/auth/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/add_encounter', methods=['GET'])
@login_required
def add_encounter():
    user = get_current_user()
    if user.role.name == 'admin':
        schemes = list(InsuranceSchemeServices.get_all())
    else:
        schemes = get_current_user().facility.scheme
        if len(schemes) == 1:
            scheme_id = schemes[0].id
            return redirect(url_for('add_scheme_encounter', scheme_id=scheme_id))

    return render_template('add_encounter.html',
                           title='Select Insurance Scheme',
                           schemes=schemes)


@app.route("/add_encounter/amchis/child_health", methods=['GET', 'POST'])
@login_required
@scheme_access_required(SchemeEnum.AMCHIS)
def add_child_health_encounter():
    scheme = InsuranceSchemeServices.get_scheme_by_enum(SchemeEnum.AMCHIS)
    if not (orin := request.args.get("orin")):
        flash("Please provide Mother ORIN for AMCHIS CHILD HEALTH Encounter ", "error")
        return redirect(url_for('add_amchis_encounter'))

    form = ChildHealthEncounterForm()
    if request.method == 'GET':
        form.policy_number.data = orin

    disease_choices = [('0', 'Select a disease')] + sorted([(dis.id, str(dis.name).title())
                                                            for dis in DiseaseServices.get_all()], key=lambda x: x[1])
    service_choices = [('0', "Select a  Service")] + sorted([(srv.id, str(srv.name).title())
                                                            for srv in ServiceServices.get_all()], key = lambda x: x[1])
    for subfield in form.diseases:
        subfield.choices = disease_choices
    for subfield in form.services:
        subfield.choices = service_choices

    if form.validate_on_submit():
        try:
            res = form_to_dict(form, Encounter)
            final_outcome = form.death_type.data if form.outcome.data == -1 else form.outcome.data
            res['outcome'] = final_outcome
            res['guardian_name'] = form.guardian_name.data
            res['dob'] = form.dob.data
            res['created_by'] = get_current_user().id
            res['age_group'] = form.age_group.data if form.age.data == 0 else get_age_group(form.age.data)
            diseases = [disease.data for disease in form.diseases]
            services = [service.data for service in form.services]
            res['diseases_id'] = diseases
            res['services_id'] = services
            user = get_current_user()
            res['created_by'] = user.id
            if user.role.name != 'admin':
                res['facility_id'] = user.facility.id
            else:
                res['facility_id'] = form.facility.data

            res['scheme'] = scheme.id
            EncounterServices.create_child_health_encounter(**res)
            flash('Encounter has successfully been added', 'success')
            return redirect(url_for('add_encounter'))
        except (InvalidReferenceError, ValidationError, ServiceError, MissingError) as e:
            flash(str(e), 'error')
    elif form.errors:
        flash("Encounter not submitted. Please check and correct errors before submitting", 'error')

    return render_template('add_child_health_encounter.html',
                           title = "Add Child Health Encounter",
                           insurance_scheme = scheme,
                           disease_choices = disease_choices[1:],
                           service_choices = service_choices[1:],
                           form = form)

@app.route("/add_encounter/amchis/delivery", methods = ['GET', 'POST'])
@login_required
@scheme_access_required(SchemeEnum.AMCHIS)
def add_delivery_encounter():
    scheme = InsuranceSchemeServices.get_scheme_by_enum(SchemeEnum.AMCHIS)
    registry_val = None
    if not (orin := request.args.get('orin')):
        flash("No ORIN provided for AMCHIS Scheme", 'error')
        return redirect(url_for('add_amchis_encounter'))
    try:
        registry_val = EncounterServices.get_anc_record_by_registry(orin)
    except MissingError:
        flash("Encounter not allowed: User is not registered for ANC.", "error")
        return redirect(url_for("add_amchis_encounter"))
    user = get_current_user()

    form = DeliveryEncounterForm()
    if registry_val and request.method == 'GET':
        form.policy_number.data = registry_val.orin
        form.client_name.data = registry_val.client_name
        form.kia_date.data = registry_val.kia_date
        form.booking_date.data = registry_val.booking_date
        form.place_of_issue.data = registry_val.place_of_issue
        form.hospital_number.data = registry_val.hospital_number
        form.address.data = registry_val.address
        form.parity.data = registry_val.parity
        form.lmp.data = registry_val.lmp
        form.anc_count.data = registry_val.anc_count
        form.age.data = registry_val.age
        form.age_group.data = registry_val.age_group
        form.nin.data = registry_val.nin
        form.phone_number.data = registry_val.phone_number
        form.expected_delivery_date.data = registry_val.expected_delivery_date
        form.gestational_age.data = calculate_gestational_age(registry_val.lmp)
    else:
        form.policy_number.data = orin

    if form.validate_on_submit():
        try:
            res = form_to_dict(form, Encounter)

            final_outcome = form.death_type.data if form.outcome.data == -1 else form.outcome.data
            res['outcome'] = final_outcome
            res['scheme'] = scheme.id

            if user.role.name.lower() != 'admin':
                res['facility_id'] = get_current_user().facility.id
            else:
                res['facility_id'] = form.facility.data

            res['anc_count'] = registry_val.anc_count if registry_val else 1
            res['mode_of_delivery'] = form.mode_of_delivery.data
            res['baby_details'] = form.babies_data
            res['age_group'] = form.age_group.data if form.age.data == 0 else get_age_group(form.age.data)
            res['anc_id'] = registry_val.id
            res['created_by'] = get_current_user().id
            res['gender'] = 'F'

            # print(res)
            EncounterServices.create_delivery_encounter(**res)
            flash("Delivery Encounter added successfully", "success")
            return redirect(url_for("add_encounter"))

        except (ServiceError, ValidationError, MissingError) as e:
            flash(str(e), "error")
            return redirect(url_for('add_amchis_encounter'))

    elif form.errors:
        flash("Encounter not submitted. Please check and correct errors before submitting", 'error')

    baby_outcome_choices = [(b.value, b.value) for b in BabyOutcome]

    return render_template("add_delivery_encounter.html",
                           title = "Add Delivery Encounter",
                           form = form,
                           insurance_scheme = scheme,
                           baby_outcome_choices = baby_outcome_choices)


@app.route('/add_encounter/amchis/anc', methods = ['GET', 'POST'])
@login_required
@scheme_access_required(SchemeEnum.AMCHIS)
def add_anc_encounter():
    scheme = InsuranceSchemeServices.get_scheme_by_enum(SchemeEnum.AMCHIS)
    registry_val = None
    if not (orin := request.args.get('orin')):
        flash("No ORIN provided for AMCHIS Scheme", 'error')
        return redirect('add_encounter')

    try:
        registry_val = EncounterServices.get_anc_record_by_registry(orin = orin)
    except MissingError:
        pass
    form = ANCEncounterForm()
    user = get_current_user()

    if registry_val and request.method == 'GET':
        form.policy_number.data = registry_val.orin
        form.client_name.data = registry_val.client_name
        form.kia_date.data = registry_val.kia_date
        form.booking_date.data = registry_val.booking_date
        form.place_of_issue.data = registry_val.place_of_issue
        form.hospital_number.data = registry_val.hospital_number
        form.address.data = registry_val.address
        form.parity.data = registry_val.parity
        form.lmp.data = registry_val.lmp
        form.nin.data = registry_val.nin
        form.phone_number.data = registry_val.phone_number
        form.expected_delivery_date.data = registry_val.expected_delivery_date
        form.gestational_age.data = calculate_gestational_age(registry_val.lmp)
    else:
        form.policy_number.data = orin

    if form.validate_on_submit():
        try:
            res =  form_to_dict(form, Encounter)
            res.update(form_to_dict(form, ANCRegistry))
            final_outcome = form.death_type.data if form.outcome.data == -1 else form.outcome.data
            res['outcome'] = final_outcome
            res['scheme'] = scheme.id
            res['age_group'] = form.age_group.data if form.age.data == 0 else get_age_group(form.age.data)
            res['created_by'] = get_current_user().id
            res['anc_count'] = (
                registry_val.anc_count + 1 if registry_val
                else 1)
            res['gender'] = 'F'
            if user.role.name.lower() != 'admin':
                res['facility_id'] = get_current_user().facility.id
            else:
                res['facility_id'] = form.facility.data

            EncounterServices.create_anc_encounter(**res)
            flash("ANC Encounter added successfully", 'success')
            return redirect(url_for("add_encounter"))

        except (ServiceError, ValidationError, MissingError) as e:
            flash(str(e), "error")
    elif form.errors:
        # print(form.errors)
        flash("Encounter not submitted. Please check and correct errors before submitting", 'error')
    else:
        print("Not doing anything")

    return render_template("add_anc_encounter.html",
                            title = "Add ANC Encounter",
                            insurance_scheme = scheme,
                            form = form)

@app.route('/add_encounter/amchis', methods=['GET', 'POST'])
@login_required
@scheme_access_required(SchemeEnum.AMCHIS)
def add_amchis_encounter():
    form: FlaskForm = EncTypeForm()
    if form.validate_on_submit():
        orin = form.orin.data
        enc_type = form.enc_type.data
        if enc_type.lower() == 'anc':
            return redirect(url_for('add_anc_encounter', orin = orin))
        elif enc_type.lower() == 'delivery':
            return redirect(url_for('add_delivery_encounter', orin = orin))
        elif enc_type.lower() == 'child health':
            return redirect(url_for('add_child_health_encounter', orin = orin))
        else:
            flash("Invalid Encounter Type Selection", 'error')
            return redirect(url_for('add_encounter'))

    return render_template('amchis_encounter.html',
                           form = form,
                           title = "Select AMCHIS ENCOUNTER")


@app.route('/add_encounter/<int:scheme_id>', methods=['GET', 'POST'])
@login_required
def add_scheme_encounter(scheme_id) -> Any:
    try:
        insurance_scheme = InsuranceSchemeServices.get_by_id(
            scheme_id)  # check if scheme exist
        user = get_current_user()

        if user.role.name != 'admin':
            schemes = get_current_user().facility.scheme
            if scheme_id not in (sc.id for sc in schemes):
                raise ValidationError(
                    f"Your facility is not under the insurance scheme: {insurance_scheme.scheme_name}")
    except MissingError:
        flash("Invalid Insurance Scheme Selected", "error")
        return redirect(url_for('add_encounter'))
    except ValidationError as e:
        flash(str(e), "error")
        return redirect(url_for('add_encounter'))

    from app.route_handler import Handler
    if ((func := Handler.get_handler(insurance_scheme))):
        return func()

    form: AddEncounterForm = AddEncounterForm()
    disease_choices = [('0', 'Select a disease')] + sorted([(dis.id, str(dis.name).title())
                                                            for dis in DiseaseServices.get_all()], key=lambda x: x[1])
    service_choices = [('0', "Select a  Service")] + sorted([(srv.id, str(srv.name).title())
                                                            for srv in ServiceServices.get_all()], key = lambda x: x[1])
    for subfield in form.diseases:
        subfield.choices = disease_choices

    for subfield in form.services:
        subfield.choices = service_choices

    if form.validate_on_submit():
        try:
            res = form_to_dict(form, Encounter)
            final_outcome = form.death_type.data if form.outcome.data == -1 else form.outcome.data
            res['outcome'] = final_outcome
            diseases = [disease.data for disease in form.diseases]
            services = [service.data for service in form.services]
            res['diseases_id'] = diseases
            res['services_id'] = services
            user = get_current_user()
            res['created_by'] = user.id
            res['age_group'] = form.age_group.data if form.age.data == 0 else get_age_group(form.age.data)
            if user.role.name != 'admin':
                res['facility_id'] = user.facility.id
            else:
                if not form.facility.data:
                    raise ValidationError("Admin user has to select a facility", "error")
                res['facility_id'] = form.facility.data

            res['scheme'] = scheme_id
            EncounterServices.create_encounter(**res)
            flash('Encounter has successfully been added', 'success')
            return redirect(url_for('add_encounter'))
        except (InvalidReferenceError, ValidationError, ServiceError, MissingError) as e:
            flash(str(e), 'error')
        # except:
            # abort(500)
    elif form.errors:
        flash("Encounter not submitted. Please check and correct errors before submitting", 'error')

    if request.method == 'GET':
        form.date.data = date.today()

    return render_template('add_scheme_encounter.html',
                           disease_choices=disease_choices[1:],
                           service_choices = service_choices[1:],
                           insurance_scheme=insurance_scheme,
                           form=form,
                           title='Add Encounter')

@app.route('/admin/claims', methods=['GET'])
@admin_required
def claims():
    return render_template('claims.html', title='Claims Management')

@app.route('/admin/facilities', methods=['GET', 'POST'])
@admin_required
def facilities() -> Any:

    facility_form = AddFacilityForm()
    if facility_form.validate_on_submit():
        res = form_to_dict(facility_form, Facility)

        res['scheme'] = facility_form.scheme.data
        res['lga'] = facility_form.lga.data
        try:
            FacilityServices.create_facility(**res)
            flash("You have successfully created a new facility", 'success')
            return redirect(url_for('facilities'))
        except (ValidationError, DuplicateError) as e:
            flash(str(e), 'error')

    filter_form = FacilityFilterForm(request.args)
    params = build_filter(filter_form,
                          ['facility_type', 'ownership', 'scheme', 'lga'],
                          Params(),
                          facility_filter_config)

    if (request.args.get('download') == 'true'):
        output =  DownloadServices.download_facilities_sheet(params)
        return send_file(output,
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        download_name = f"master_facility_sheet.xlsx",
                        as_attachment=True)

    if (limit := request.args.get('limit')):
        params = params.set_limit(int(limit))

    primary_total = FacilityServices.get_total(Params().where(Facility, 'facility_type', '=', 'Primary'))
    secondary_total = FacilityServices.get_total(Params().where(Facility, 'facility_type', '=', 'Secondary'))
    tertiary_total = FacilityServices.get_total(Params().where(Facility, 'facility_type', '=', 'Tertiary'))

    facility_total = FacilityServices.get_total(Params())

    page = int(request.args.get('page', 1))
    facility_list = list(FacilityServices.list_row_by_page(page, params=params))

    # Pagination Logic
    res_args = {**request.args}
    res_args['page'] = page + 1
    next_url = url_for('facilities', **res_args) if FacilityServices.has_next_page(page, params=params) else None

    res_args['page'] = page - 1
    prev_url = url_for('facilities', **res_args) if page > 1 else None

    return render_template('facilities.html',
                           title='Facilities',
                           prev_url=prev_url,
                           next_url=next_url,
                           # Pass GLOBAL totals to cards
                           facility_total=facility_total,
                           primary_total=primary_total,
                           secondary_total=secondary_total,
                           tertiary_total=tertiary_total,
                           # Pass Forms & List
                           facility_form=facility_form,
                           filter_form=filter_form,
                           facility_list=facility_list,
                           current_page=page,
                           # Pass current_type for CSS Active State
                           current_type= filter_form.facility_type.data)


@app.route('/admin/facilities/edit/<int:pid>', methods=['GET', 'POST'])
@admin_required
def edit_facilities(pid: int) -> Any:
    try:
        facility = FacilityServices.get_by_id(pid)
    except MissingError as e:
        flash(str(e), 'error')
        return redirect(url_for('facilities'))
    # except:
        # abort(500)

    form: FlaskForm = EditFacilityForm(obj=facility)
    others = [(sc.id, sc.scheme_name)
              for sc in InsuranceSchemeServices.get_all()]
    form.scheme.choices = others
    current_scheme = FacilityServices.get_current_scheme(pid)
    form.scheme.data = current_scheme

    if form.validate_on_submit():
        try:
            form.populate_obj(facility)
            facility.local_government = form.lga.data
            FacilityServices.update_facility(facility, form.scheme.data)
            flash("You have successfully added a new facility", 'success')
            return redirect(url_for('facilities'))
        except (DuplicateError, ValidationError) as e:
            flash(str(e), 'error')
        # except:
            # abort(500)

    return render_template('edit_facilities.html',
                           facility=facility,
                           form=form,
                           title='Edit Facility')


@app.route('/admin/facilities/view/<int:pid>', methods=['GET', 'POST'])
@admin_required
def view_facilities(pid: int) -> Any:
    try:
        facility = FacilityServices.get_view_by_id(pid)
        # print(pid, facility)
    except MissingError as e:
        flash(str(e), 'error')
        return redirect(url_for('facilities'))
    # except:
        # abort(500)

    user_form: AddUserForm = AddUserForm()
    user_form.facility_id.choices = ([('0', 'Select a facility')] +
                                     sorted([(fac.id, fac.name.title())
                                             for fac in FacilityServices.get_all()], key=lambda x: x[1]))
    user_form.facility_id.data = pid

    if user_form.validate_on_submit():
        try:
            UserServices.create_user(username=user_form.username.data,
                                     facility_id=pid,
                                     password=user_form.password.data)
            return redirect(url_for('view_facilities', pid=pid))
        except (DuplicateError, InvalidReferenceError, ValidationError) as e:
            flash(str(e), "error")

    today = datetime.now().date()
    first_month_day = today.replace(day=1)


    user_count = UserServices.get_total(params= Params().where(User, 'facility_id', '=', pid))

    filters = Params()
    filters = (filters.where(Encounter, 'facility_id', '=', pid)
            .where(Encounter, 'date', '>=', first_month_day)
            .where(Encounter, 'date', '<=', today))

    month_encounter_count = EncounterServices.get_total(params = filters)
    filters = filters.set_limit(10)
    filters = filters.sort(Encounter, 'date', 'DESC')
    recent_encounters = list(EncounterServices.get_all(params = filters))


    page = int(request.args.get('user_page', 1))
    user_list = list(UserServices.list_row_by_page(page=page,
                                                   params=Params().where(User, 'facility_id', '=', pid)))

    next_url = url_for('view_facilities', pid=pid, user_page=page +
                       1) if UserServices.has_next_page(page,
                                                        params=Params().where(User, 'facility_id', '=', pid)) else None
    prev_url = url_for('view_facilities', pid=pid,
                       user_page=page - 1) if page > 1 else None
    # print(facility)


    return render_template('view_facilities.html',
                           title = f"Viewing Facility: {facility.name}",
                           month_encounter_count=month_encounter_count,
                           facility=facility,
                           recent_encounters=recent_encounters,
                           user_list=user_list,
                           user_count=user_count,
                           user_form=user_form,
                           next_url=next_url,
                           prev_url=prev_url
                           )

@app.route('/admin/services', methods=['GET'])
@admin_required
def services():
    page = int(request.args.get('page', 1))
    filters = Params()
    category_list = list(ServiceCategoryServices.get_all())
    if category := request.args.get('category'):
        filters = filters.where(Service, 'category_id', '=', category)

    if request.args.get('download') == 'true':
        output = DownloadServices.download_services_sheet(params=filters)
        return send_file(output,
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        download_name = f"master_services_sheet.xlsx",
                        as_attachment=True)

    service_list = list(ServiceServices.list_row_by_page(page, params=filters))
    # Get filtered count for pagination
    filtered_services = ServiceServices.get_total(params=filters)

    # Get total unfiltered count for the card
    total_services = ServiceServices.get_total()

    total_categories = ServiceCategoryServices.get_total()

    # Get active category name
    active_category_name = None
    if category:
        active_cat = next((cat for cat in category_list if str(cat.id) == str(category)), None)
        if active_cat:
            active_category_name = active_cat.name

    res = {** request.args}
    if category:
        res['category'] = category

    res['page'] = page+1
    next_url = url_for('services', **res) if ServiceServices.has_next_page(page=page,
                                                                           params=filters) else None
    res['page'] = page - 1
    prev_url = url_for('services', **res) if page > 1 else None

    return render_template('services.html',
                           title='Manage Services',
                           service_list= service_list,
                           total_services= total_services,
                           filtered_services=filtered_services,
                           total_categories=total_categories,
                           current_page=page,
                           total_pages=20,
                           per_page=Config.ADMIN_PAGE_PAGINATION,
                           category_list=category_list,
                           active_category=int(category) if category else 0,
                           active_category_name=active_category_name,
                           next_url=next_url,
                           prev_url=prev_url)

@app.route('/admin/services/category/add', methods=['GET', 'POST'])
@admin_required
def add_service_category():
    form = AddCategoryForm()
    if form.validate_on_submit():
        try:
            ServiceCategoryServices.create_category(
                category_name=form.category_name.data)

            flash('Service category added successfully', 'success')
            return redirect(url_for('services'))
        except DuplicateError as e:
            flash(str(e), 'error')
        # except Exception:
            # abort(500)
    return render_template('add_service_category.html', title='Add Service Category', form=form)


@app.route('/admin/services/add', methods=['GET', 'POST'])
@admin_required
def add_service():
    form = AddServiceForm()
    form.category_id.choices = ([('0', 'Select a Category')] +
                                sorted([(cat.id, cat.name.title())
                                        for cat in ServiceCategoryServices.get_all()], key=lambda x: x[1]))

    if form.validate_on_submit():
        try:
            ServiceServices.create_service(
                name=form.name.data, category_id=form.category_id.data)
            flash('New disease added successfully', 'success')

            return redirect(url_for('services'))
        except (DuplicateError, InvalidReferenceError) as e:
            flash(str(e), 'error')
    return render_template('add_service.html',
                            title='Add Service',
                            service = None,
                            form=form)

@app.route('/admin/services/edit/<int:service_id>', methods=['GET', 'POST'])
@admin_required
def edit_service(service_id: int):
    try:
        service = ServiceServices.get_by_id(service_id)
    except MissingError:
        abort(404)

    form = EditDiseaseForm(obj=service)  # Pre-populate form with existing data
    form.category_id.choices = [('0', 'Select a Category')] + sorted(
        [(cat.id, cat.name.title())
         for cat in ServiceCategoryServices.get_all()],
        key=lambda x: x[1])

    if form.validate_on_submit():
        try:
            form.populate_obj(service)
            ServiceServices.update_service(service)
            flash(f"Service '{service.name}' has been updated.", 'success')
            return redirect(url_for('services'))
        except (DuplicateError, InvalidReferenceError) as e:
            flash(str(e), 'error')

    return render_template('add_service.html',
                            title=f"Edit Service: {service.name}",
                            form=form,
                            service=service)


@app.route('/admin/diseases', methods=['GET'])
@admin_required
def diseases():
    page = int(request.args.get('page', 1))
    filters = Params()
    category_list = list(DiseaseCategoryServices.get_all())

    if category := request.args.get('category'):
        filters = filters.where(Disease, 'category_id', '=', category)

    if request.args.get('download') == 'true':
        output = DownloadServices.download_diseases_sheet(params=filters)
        return send_file(output,
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        download_name = f"master_diseases_sheet.xlsx",
                        as_attachment=True)

    disease_list = list(DiseaseServices.list_row_by_page(page, params = filters))
    filtered_diseases = DiseaseServices.get_total(params=filters)

    total_diseases = DiseaseServices.get_total()

    total_categories = DiseaseCategoryServices.get_total()
    total_category_diseases = DiseaseServices.get_total(params=filters)

    active_category_name = None
    if category:
        active_cat = next((cat for cat in category_list if str(cat.id) == str(category)), None)
        if active_cat:
            active_category_name = active_cat.category_name

    # Get active category name
    active_category_name = None
    if category:
        active_cat = next((cat for cat in category_list if str(cat.id) == str(category)), None)
        if active_cat:
            active_category_name = active_cat.category_name

    res = {** request.args}
    if category:
        res['category'] = category

    res['page'] = page+1
    next_url = url_for('diseases', **res) if DiseaseServices.has_next_page(page=page,
                                                                           params=filters) else None
    res['page'] = page - 1
    prev_url = url_for('diseases', **res) if page > 1 else None

    return render_template('diseases.html',
                           title='Manage Diseases',
                           disease_list=disease_list,
                           total_diseases=total_diseases,
                           filtered_diseases=filtered_diseases,
                           total_categories=total_categories,
                           total_category_diseases = total_category_diseases,
                           current_page=page,
                           total_pages=20,
                           per_page=Config.ADMIN_PAGE_PAGINATION,
                           category_list=category_list,
                           active_category=int(category) if category else 0,
                           active_category_name=active_category_name,
                           next_url=next_url,
                           prev_url=prev_url)


@app.route('/admin/diseases/category/add', methods=['GET', 'POST'])
@admin_required
def add_category():
    form = AddCategoryForm()
    if form.validate_on_submit():
        try:
            DiseaseCategoryServices.create_category(
                category_name=form.category_name.data)
            flash('Disease category added successfully', 'success')
            return redirect(url_for('diseases'))
        except DuplicateError as e:
            flash(str(e), 'error')
        # except Exception:
            # abort(500)
    return render_template('add_disease_category.html', title='Add Disease Category', form=form)


@app.route('/admin/diseases/add', methods=['GET', 'POST'])
@admin_required
def add_disease():
    form = AddDiseaseForm()
    form.category_id.choices = ([('0', 'Select a Category')] +
                                sorted([(cat.id, cat.category_name.title())
                                        for cat in DiseaseCategoryServices.get_all()], key=lambda x: x[1]))

    if form.validate_on_submit():
        try:
            DiseaseServices.create_disease(
                name=form.name.data, category_id=form.category_id.data)
            flash('New disease added successfully', 'success')
            return redirect(url_for('diseases'))
        except (DuplicateError, InvalidReferenceError) as e:
            flash(str(e), 'error')
        # except Exception:
            # abort(500)
    return render_template('add_disease.html', title='Add Disease', form=form)


@app.route('/admin/diseases/edit/<int:disease_id>', methods=['GET', 'POST'])
@admin_required
def edit_disease(disease_id: int):
    try:
        disease = DiseaseServices.get_by_id(disease_id)
    except MissingError:
        abort(404)

    form = EditDiseaseForm(obj=disease)  # Pre-populate form with existing data
    form.category_id.choices = [('0', 'Select a Category')] + sorted(
        [(cat.id, cat.category_name.title())
         for cat in DiseaseCategoryServices.get_all()],
        key=lambda x: x[1])

    if form.validate_on_submit():
        try:
            form.populate_obj(disease)
            DiseaseServices.update_disease(disease)
            flash(f"Disease '{disease.name}' has been updated.", 'success')
            return redirect(url_for('diseases'))
        except (DuplicateError, InvalidReferenceError) as e:
            flash(str(e), 'error')
        # except Exception:
            # abort(500)
    return render_template('add_disease.html', title=f"Edit Disease: {disease.name}", form=form, disease=disease)


@app.route('/admin/users', methods=['GET', 'POST'])
@admin_required
def users():
    page = int(request.args.get('page', 1))
    user_form: AddUserForm = AddUserForm()

    if not user_form.facility_id.choices:
        user_form.facility_id.choices = [('0', 'Select a facility')] + [(
            fac.id, fac.name.title()) for fac in FacilityServices.get_all()]

    if user_form.validate_on_submit():
        res = form_to_dict(user_form, User)
        res['password'] = user_form.password.data
        try:
            user = UserServices.create_user(**res)
            facility: Facility = FacilityServices.get_by_id(user.facility_id)
            flash(f'{user.username} added to {facility.name} facility',
                  'success')
            return redirect(url_for('users'))
        except (DuplicateError, InvalidReferenceError) as e:
            flash(str(e), 'error')

    user_list = list(UserServices.list_row_by_page(page))
    total_user = UserServices.get_total()
    next_url = (url_for('users', page=page + 1)
                if UserServices.has_next_page(page) else None)
    prev_url = None if page == 1 else url_for('users', page=page - 1)

    return render_template('users.html',
                           title="Manage User",
                           user_form=user_form,
                           user_list=user_list,
                           total_user=total_user,
                           next_url=next_url,
                           prev_url=prev_url)

@app.route('/admin/users/edit/<int:user_id>', methods=['GET', 'POST'])
@admin_required
def edit_user(user_id: int):
    try:
        user = UserServices.get_by_id(user_id)
    except MissingError:
        flash('User not found in database')
        return redirect(url_for('users'))
    form = EditUserForm()

    if form.validate_on_submit():
        try:
            if user.username != form.username.data:
                user.username = form.username.data
                UserServices.update_user(user)
            if form.password.data:
                UserServices.update_user_password(user, form.password.data)
            flash('User updated successfully', 'success')
            return redirect(url_for('users'))
        except DuplicateError as e:
            flash(str(e), 'error')

    if request.method == 'GET':
        form.username.data = user.username
    delete_form = DeleteUserForm()
    return render_template('edit_user.html',
                           title=f"Edit user {user.username}",
                           user=user,
                           edit_form=form,
                           delete_form=delete_form)

@app.route('/admin/user/delete/<int:user_id>', methods=['POST'])
@admin_required
def delete_user(user_id: int):
    user = UserServices.get_by_id(user_id)
    delete_form = DeleteUserForm()

    if delete_form.validate_on_submit():
        UserServices.delete_user(user)
        flash(f"User '{user.username}' has been deleted.", 'success')
    else:
        flash("CSRF token invalid. Could not delete user.", 'error')

    return redirect(url_for('users'))

@app.route('/encounters')
@login_required
def encounters():
    user = get_current_user()
    page = int(request.args.get('page', 1))
    filter_form = EncounterFilterForm(request.args)
    user = get_current_user()
    if user.facility:
        filter_form.scheme_id.choices = [(s.id, s.scheme_name) for s in user.facility.scheme]

    if request.args.get('download') == 'true':
        filters = build_filter(filter_form, ['period', 'scheme_id', 'outcome', 'facility_id', 'age_group'], Params(), download_encounter_filter_config)
        res = DownloadServices.download_encounter_sheet(params=filters)
        return send_file(
            res,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=secure_filename('master_encounter_report.xlsx')
        )

    filters = build_filter(filter_form, ['period', 'scheme_id', 'outcome', 'facility_id', 'age_group'], Params(), encounter_filter_config)

    encounter_list = list(EncounterServices.list_row_by_page(page,
                                                             params=filters))

    pagination_args = {**request.args, "page": page + 1}
    next_url = (url_for('encounters', **pagination_args)
                if EncounterServices.has_next_page(page, params=filters) else None)

    pagination_args['page'] = page - 1
    prev_url = None if page == 1 else url_for('encounters', **pagination_args)
    return render_template('encounters.html',
                           title="Encounter List",
                           encounter_list=encounter_list,
                           filter_form=filter_form,
                           next_url=next_url,
                           prev_url=prev_url
                           )

@app.route('/encounters/view/<int:pid>')
@login_required
def view_encounter(pid: int):
    try:
        encounter_view = EncounterServices.get_view_by_id(pid)
        user = get_current_user()
        if user.role.name != 'admin' and user.facility.id != encounter_view.facility.id:
            raise ValidationError("Encounter not registered to your facility")
    except (MissingError, ValidationError) as e:
        flash(str(e), "error")
        return redirect(url_for("encounters"))
    return render_template('view_encounter.html',
                           title=f"Encounter Details: {encounter_view.client_name}",
                           encounter=encounter_view)


@app.route('/dashboard/overview')
@admin_required
def admin_overview():

    start_date, end_date = parse_date()
    g.start_date = start_date
    g.end_date = end_date
    form = DashboardFilterForm(request.args)
    if not form.validate():
        flash("Invalid Filter Parameters", 'error')
        return redirect(url_for('admin_overview'))

    all_filter = build_filter(form, ['period', 'scheme_id', 'gender'] )
    without_date_filter = build_filter(form, [ 'scheme_id', 'gender'] )

    total_facilities = DashboardServices.get_active_encounter_facility(all_filter)
    total_encounter = DashboardServices.get_total_encounters(params = without_date_filter,
                                                    start_date = g.start_date, end_date = g.end_date)
    total_death = DashboardServices.get_total_death_outcome(without_date_filter, g.start_date, g.end_date)
    facilities_summary = DashboardServices.get_top_facilities_summaries(all_filter, g.start_date, g.end_date)
    total_utilization = DashboardServices.get_total_utilization(without_date_filter, g.start_date, g.end_date)
    encounter_scheme_grouped = DashboardServices.total_encounter_by_scheme_grouped(all_filter, g.start_date,
                                                                                   g.end_date)

    utilization_scheme_grouped = DashboardServices.total_utilization_by_scheme_grouped(all_filter, g.start_date,
                                                                                       g.end_date)
    mortality_scheme_grouped = DashboardServices.total_mortality_by_scheme_grouped(all_filter, g.start_date,
                                                                                       g.end_date)
    # print(f'total_facilities: {total_facilities}, total_death: {total_death} total_encounter: {total_encounter}')

    return render_template(
        'dashboard_overview.html',
        title = 'Dashboard',
        total_facilities = total_facilities,
        total_encounter = total_encounter,
        total_death = total_death,
        facilities_summary = facilities_summary,
        total_utilization = total_utilization,
        encounter_scheme_grouped = encounter_scheme_grouped,
        utilization_scheme_grouped = utilization_scheme_grouped,
        mortality_scheme_grouped = mortality_scheme_grouped,
        form = form,
        start_date = g.start_date,
        end_date = g.end_date
    )


@app.route('/dashboard/utilization')
@admin_required
def admin_utilization():
    start_date, end_date = parse_date()
    g.start_date = start_date
    g.end_date = end_date
    form = AdminDashboardFilterForm(request.args)
    if not form.validate():
        flash("Invalid Filter Parameters", "error")
        return redirect(url_for('admin_utilization'))

    base_list = ['period']
    all_filters = build_filter(form, base_list + ['scheme_id' , 'lga' , 'gender' ,'facility_id'])
    without_facility_filters = build_filter(form, base_list + ['scheme_id' , 'lga' , 'gender'])
    without_gender_filters = build_filter(form, base_list + ['scheme_id' , 'lga' , 'facility_id'])
    without_lgas_filters = build_filter(form, base_list + ['scheme_id' , 'gender' , 'facility_id'])
    without_scheme_filters = build_filter(form, base_list + ['lga' , 'gender' , 'facility_id'])
    without_date_filters = build_filter(form, ['lga', 'gender', 'scheme_id', 'facility_id'])

    utilization_per_scheme = DashboardServices.get_utilization_per_scheme(all_filters)
    utilization_per_lga = DashboardServices.utilization_distribution_across_lga(without_lgas_filters)
    top_utilized_items = DashboardServices.top_utilized_items(all_filters)
    average_daily_utilization = DashboardServices.get_average_utilization_per_day(without_date_filters, g.start_date, g.end_date)
    total_utilization = DashboardServices.get_total_utilization(without_date_filters, g.start_date, g.end_date)
    utilization_age_distribution = DashboardServices.utilization_age_group_distribution(all_filters)
    top_utilized_facilities = DashboardServices.get_top_utilization_facilities(all_filters)
    service_utilization_rate =  DashboardServices.get_service_utilization_rate(without_date_filters, g.start_date, g.end_date)
    utilization_trend = DashboardServices.get_utilization_trend(without_date_filters, g.start_date, g.end_date)

    return render_template(
        'dashboard_utilization.html',
        title = 'Dashboard',
        total_utilization = total_utilization,
        utilization_per_scheme = utilization_per_scheme,
        utilization_age_distribution = utilization_age_distribution,
        top_utilized_items = top_utilized_items,
        average_daily_utilization = average_daily_utilization,
        utilization_trend = utilization_trend,
        top_utilized_facilites = top_utilized_facilities,
        service_utilization_rate = service_utilization_rate,
        utilization_per_lga = utilization_per_lga,
        start_date = g.start_date,
        end_date = g.end_date,
        form = form
        )


@app.route('/dashboard/encounters')
@admin_required
def admin_encounters():
    start_date, end_date = parse_date()
    g.start_date = start_date
    g.end_date = end_date

    form = AdminDashboardFilterForm(request.args)
    if not form.validate():
        flash("Invalid Filter Parameters", "error")
        return redirect(url_for('admin_encounters'))

    base_list = ['period']
    all_filters = build_filter(form, base_list + ['scheme_id' , 'lga' , 'gender' ,'facility_id'])
    without_facility_filters = build_filter(form, base_list + ['scheme_id' , 'lga' , 'gender'])
    without_gender_filters = build_filter(form, base_list + ['scheme_id' , 'lga' , 'facility_id'])
    without_lgas_filters = build_filter(form, base_list + ['scheme_id' , 'gender' , 'facility_id'])
    without_scheme_filters = build_filter(form, base_list + ['lga' , 'gender' , 'facility_id'])
    without_date_filters = build_filter(form, ['lga', 'gender', 'scheme_id', 'facility_id'])

    total_encounter = DashboardServices.get_total_encounters(without_date_filters, g.start_date, g.end_date)
    encounter_gender_distribution = DashboardServices.encounter_gender_distribution(without_gender_filters)
    encounter_age_distribution = DashboardServices.encounter_age_group_distribution(all_filters)
    encounter_per_scheme = DashboardServices.get_encounter_per_scheme(without_scheme_filters)
    treatment_outcome_distribution = DashboardServices.get_treatment_outcome_distribution(all_filters)
    top_encounter_facilities = DashboardServices.get_top_encounter_facilities(all_filters)
    average_daily_encounter = DashboardServices.get_average_encounter_per_day(all_filters, g.start_date, g.end_date)
    encounter_trend = DashboardServices.get_encounter_trend(without_date_filters, g.start_date, g.end_date)
    encounter_per_lga = DashboardServices.encounter_distribution_across_lga(without_lgas_filters)

    return render_template(
        'dashboard_encounters.html',
        title = 'Dashboard',
        total_encounter = total_encounter,
        encounter_gender_distribution = encounter_gender_distribution,
        encounter_age_distribution = encounter_age_distribution,
        treatment_outcome_distribution = treatment_outcome_distribution,
        encounter_per_scheme = encounter_per_scheme,
        top_encounter_facilities = top_encounter_facilities,
        average_daily_encounter = average_daily_encounter,
        encounter_trend = encounter_trend,
        encounter_per_lga = encounter_per_lga,
        start_date = g.start_date,
        end_date = g.end_date,
        form = form
    )

@app.route('/dashboard/mortality')
@admin_required
def admin_mortality():
    start_date, end_date = parse_date()
    g.start_date = start_date
    g.end_date = end_date

    form = AdminDashboardFilterForm(request.args)
    if not form.validate():
        flash("Invalid Filter Parameters", 'error')
        return redirect(url_for('admin_mortality'))


    base_list = ['period']
    all_filters = build_filter(form, base_list + ['scheme_id' , 'lga' , 'gender' ,'facility_id'])
    without_facility_filters = build_filter(form, base_list + ['scheme_id' , 'lga' , 'gender'])
    without_gender_filters = build_filter(form, base_list + ['scheme_id' , 'lga' , 'facility_id'])
    without_lgas_filters = build_filter(form, base_list + ['scheme_id' , 'gender' , 'facility_id'])
    without_scheme_filters = build_filter(form, base_list + ['lga' , 'gender' , 'facility_id'])
    without_date_filters = build_filter(form, ['lga', 'gender', 'scheme_id', 'facility_id'])

    mortality_type_distribution  = DashboardServices.mortality_distribution_by_type(all_filters)
    mortality_age_group_distribution = DashboardServices.mortality_distribution_by_age_group(all_filters)
    mortality_facility_distribution = DashboardServices.get_mortality_count_per_facility(without_facility_filters)
    mortality_gender_distribution = DashboardServices.get_mortality_distribution_by_gender(without_gender_filters)
    mortality_scheme_distribution = DashboardServices.get_mortality_per_scheme(without_scheme_filters)
    mortality_trend = DashboardServices.get_mortality_trend(all_filters, g.start_date, g.end_date)
    mortality_per_lga = DashboardServices.get_mortality_by_lga(without_lgas_filters)
    average_daily_mortality = DashboardServices.get_average_mortality_per_day(without_date_filters, g.start_date, g.end_date)
    top_cause = DashboardServices.get_top_cause_of_mortality(all_filters)
    total_death = DashboardServices.get_total_death_outcome(without_date_filters, start_date = g.start_date, end_date = g.end_date)
    case_fatality = DashboardServices.case_fatality(all_filters)

    return render_template(
        'dashboard_mortality.html',
        title = 'Dashboard',
        mortality_type_distribution = mortality_type_distribution,
        mortality_age_group_distribution = mortality_age_group_distribution,
        mortality_facility_distribution = mortality_facility_distribution,
        mortality_gender_distribution = mortality_gender_distribution,
        mortality_scheme_distribution = mortality_scheme_distribution,
        mortality_trend_distribution = mortality_trend,
        mortality_per_lga = mortality_per_lga,
        mortality_top_cause = top_cause,
        total_death =  total_death,
        average_daily_mortality = average_daily_mortality,
        case_fatality = case_fatality,
        form = form,
        start_date = g.start_date,
        end_date = g.end_date
    )

@app.route('/admin/reports')
@admin_required
def reports():
    facilities = list(FacilityServices.get_all())

    current_year = datetime.now().year
    year_choices = [(year, year)
                    for year in range(current_year - 5, current_year + 1)]
    month = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August',
             'September', 'October', 'November', 'December']
    month_choices = list(enumerate(month, start=1))

    return render_template('reports.html',
                           title="Generate Reports",
                           facilities=facilities,
                           year_choices=year_choices,
                           month_choices=month_choices)

def get_report_data():
    report_type = request.args.get('report_type')
    facility_id_str = request.args.get('facility_id')
    month_str = request.args.get('month')
    year_str = request.args.get('year')

    month = int(month_str) if month_str else None
    year = int(year_str) if year_str else None
    facility_id = int(facility_id_str) if facility_id_str else None

    report_data = None
    report_title = ""
    start_date = None
    facility = None

    if report_type == 'utilization':
        if not facility_id:
            raise ValidationError(
                "Please select a facility for the Utilization Report.")
        facility, start_date, report_data = ReportServices.generate_service_utilization_report(
            facility=facility_id, month=month, year=year
        )
        report_title = f"Service Utilization Report for {facility.name} "

    elif report_type == 'encounter':
        start_date, report_data = ReportServices.generate_encounter_report(
            month=month, year=year)
        report_title = "Encounter Report "

    elif report_type == 'categorization':
        start_date, report_data = ReportServices.generate_categorization_report(
            month=month, year=year)
        report_title = "Disease Categorization Report "
    elif report_type == 'nhia_encounter':
        start_date, report_data = ReportServices.generate_nhia_encounter_report(month= month, year = year)
        report_title = "NHIA Encounter Report"
    else:
        raise ValidationError("Invalid report type selected.")

    return report_title, start_date, facility, report_data


@app.route('/admin/view_report')
@admin_required
def view_report():
    try:
        report_title, start_date, facility, report_data = get_report_data()
    except (MissingError, ValidationError) as e:
        flash(str(e), 'error')
        return redirect(url_for('reports'))
    except ValueError:
        flash("Invalid value provided for month, year, or facility.", 'error')
        return redirect(url_for('reports'))
    # except Exception as e:
        # abort(500)

    header_info = []
    if report_data is not None and not report_data.empty:
        if report_data.columns.nlevels > 1:
            # Handle MultiIndex
            outer_headers = []
            for col in report_data.columns:
                if col[0] not in outer_headers:
                    outer_headers.append(col[0])

            for header in outer_headers:
                loc = report_data.columns.get_loc(header)
                colspan = 1
                if isinstance(loc, slice):
                    colspan = loc.stop - loc.start
                elif hasattr(loc, 'sum'):
                    colspan = loc.sum()

                header_info.append({'name': header, 'colspan': colspan})

    return render_template('view_report.html',
                           title=report_title,
                           report_title=report_title,
                           start_date=start_date,
                           report_data=report_data,
                           header_info=header_info)


def append_utilization_header(report_data, start_date: date, facility: Facility):
    output_buffer = io.BytesIO()
    with pd.ExcelWriter(output_buffer) as writer:
        report_data.to_excel(writer, startrow=3)
    output_buffer.seek(0)

    wb = load_workbook(output_buffer)
    ws = wb.active
    ws.merge_cells("A1:S1")
    ws['A1'].value = 'ONDO STATE CONTRIBUTORY HEALTH COMMISSION'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:S2')
    ws['A2'].value = start_date.strftime("%b-%y")
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['A2'].font = Font(bold=True)

    ws.merge_cells('A3:R3')
    ws['A3'].value = f'ORANGHIS SERVICE UTILIZATION ({facility.name})'
    ws['A3'].alignment = Alignment(horizontal="center")
    ws['A3'].font = Font(bold=True)
    ws['B4'].value = 'GROUP AGE'
    ws['B4'].font = Font(bold=True)
    ws['B4'].alignment = Alignment(horizontal="center")
    ws['B5'].value = 'SEX'
    ws['B5'].font = Font(bold=True)
    ws['B5'].alignment = Alignment(horizontal="center")
    ws['B6'].font = Font(bold=True)
    ws['B6'].value = 'DISEASES'
    ws['B6'].alignment = Alignment(horizontal="center")
    ws.merge_cells('A4:A6')
    ws['A4'].value = 'S/N'
    ws['A4'].font = Font(bold=True)
    ws['A4'].alignment = Alignment(horizontal='center', vertical='center')

    for row in ws['C4':'S6']:
        for cell in row:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
    final_output = io.BytesIO()
    ws.column_dimensions['B'].width = 65
    ws.column_dimensions['S'].width = 18
    ws.column_dimensions['A'].width = 5
    wb.save(final_output)
    return final_output


def append_encounter_header(report_data, start_date: date):
    output_buffer = io.BytesIO()
    with pd.ExcelWriter(output_buffer) as writer:
        report_data.to_excel(writer, startrow=1)
    output_buffer.seek(0)

    wb = load_workbook(output_buffer)
    ws = wb.active
    ws.merge_cells("A1:S1")
    ws['A1'].value = f"{start_date.strftime('%B').upper()} ENCOUNTER PER FACILITIES"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws['B2'].value = 'GROUP AGE'
    ws['B2'].font = Font(bold=True)
    ws['B2'].alignment = Alignment(horizontal="center")
    ws['B3'].value = 'SEX'
    ws['B3'].font = Font(bold=True)
    ws['B3'].alignment = Alignment(horizontal="center")
    ws['B4'].font = Font(bold=True)
    ws['B4'].value = 'FACILITIES'
    ws['B4'].alignment = Alignment(horizontal="center")

    for row in ws['C2':'S4']:
        for cell in row:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:A4')
    # ws.column_dimensions['S'] = 20
    # ws.column_dimensions['B'] = 25
    # ws.column_dimensions['A'] = 8
    ws['A2'].value = 'S/N'
    ws['A2'].font = Font(bold=True)
    ws['A2'].alignment = Alignment(vertical="center", horizontal='center')
    ws.column_dimensions['B'].width = 65
    ws.column_dimensions['S'].width = 18
    ws.column_dimensions['A'].width = 5
    final_output = io.BytesIO()
    wb.save(final_output)
    return final_output


def append_categorization_header(report_data, start_date: date):
    output_buffer = io.BytesIO()
    with pd.ExcelWriter(output_buffer) as writer:
        report_data.to_excel(writer, startrow=1)
    output_buffer.seek(0)

    wb = load_workbook(output_buffer)
    ws = wb.active
    ws.merge_cells("A1:J1")
    ws['A1'].value = f"{start_date.strftime('%B').upper()} DISEASE CATEGORIZATION PER FACILITIES"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A2'].value = 'S/N'
    ws['A2'].font = Font(bold=True)
    ws['A2'].alignment = Alignment(vertical="center", horizontal='center')
    ws.column_dimensions['B'].width = 65
    final_output = io.BytesIO()
    wb.save(final_output)
    return final_output


def append_nhia_encounter_header(report_data, start_date: date):
    output_buffer = io.BytesIO()
    with pd.ExcelWriter(output_buffer) as writer:
        report_data.to_excel(writer)
    output_buffer.seek(0)

    wb = load_workbook(output_buffer)
    ws = wb.active
    ws['A1'].value = 'S/N'
    for idx,row in enumerate(ws.columns):
        for cell in row:
            if idx == 0:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.alignment = Alignment(horizontal='left', vertical='center')
    autofit_columns(ws, 55)
    final_output = io.BytesIO()
    wb.save(final_output)
    return final_output

@app.route('/admin/download_report')
@admin_required
def download_report():
    try:
        report_title, start_date, facility, report_data = get_report_data()
    except (MissingError, ValidationError) as e:
        flash(str(e), 'error')
        return redirect(url_for('reports'))
    except ValueError:
        flash("Invalid value provided for month, year, or facility.", 'error')
        return redirect(url_for('reports'))
    # except Exception as e:
        # abort(500)
    report_name = f"{report_title.replace(' ', '_')}_{start_date.strftime('%B')}.xlsx"

    report_type = request.args.get('report_type')
    if report_type == 'utilization':
        output_buffer = append_utilization_header(
            report_data, start_date, facility)
    elif report_type == 'encounter':
        output_buffer = append_encounter_header(report_data, start_date)
    elif report_type == 'categorization':
        output_buffer = append_categorization_header(report_data, start_date)
    elif report_type == "nhia_encounter":
        output_buffer =  append_nhia_encounter_header(report_data, start_date)
    else:
        flash("Invalid report type", "error")
        return redirect(url_for("view_report"))

    output_buffer.seek(0)
    return send_file(
        output_buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=secure_filename(report_name)
    )

@app.route('/admin/analytic_query')
@admin_required
def analytic_query():
    return render_template('analytic_query.html', title="Analytic Query")

@app.post('/admin/chat')
@admin_required
def chat():
    conversation = request.form.get('conversation_history', '')
    user_input = request.form.get('user_input', '')
    try:
        res = json.loads(conversation)
    except json.JSONDecodeError:
        res = []

    def generate():
        chatsession = ChatServices()
        for  text_chunk in chatsession.generate_response(user_input, res):
            yield f"data: {text_chunk}\n\n"
        yield "event: stop\n\n"

    content = stream_with_context(generate())
    response = Response(content, mimetype='text/event-stream')
    response.headers['cache-control'] = 'no-cache'
    response.headers['X-Accel-Buffering'] = 'no'
    response.headers['Connection'] = 'keep-alive'
    response.headers['Transfer-Encoding'] = 'chunked'
    return response


@app.route('/admin/upload_excel', methods=['GET', 'POST'])
@admin_required
def upload_excel():
    import calendar
    form = ExcelUploadForm()
    facility_list = [('0', 'Select Facility')] + [(facility.id, facility.name)
                                                  for facility in sorted(FacilityServices.get_all(), key=lambda x: x.name)]
    month_list = [('0', 'Select Month')] + \
        [(i, calendar.month_name[i]) for i in range(1, 13)]
    form.facility_id.choices = facility_list
    form.month.choices = month_list
    # if form.validate_on_submit():
    # file_data = form.facility.data
    # month = form.month
    # facility_id = form.facility_id.data
    # user = get_current_user()

    return render_template('upload_excel.html', title='Upload Encounter Sheet', form=form)
