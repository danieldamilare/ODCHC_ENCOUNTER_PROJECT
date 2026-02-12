import io
import sqlite3
import pandas as pd
from typing import Dict, Callable
from openpyxl import load_workbook
from openpyxl.styles import Font

from app.db import get_db
from app.utils import autofit_columns
from app.filter_parser import Params, FilterParser
from app.models import (
    Facility, FacilityScheme, InsuranceScheme,
    Service, ServiceCategory, Disease, DiseaseCategory
)

from .base import BaseServices

class DownloadServices(BaseServices):

    @classmethod
    def build_dataframe_buffer(cls, query: str,
                               params: Params,
                               model_map: Dict,
                               row_processor: Callable[[sqlite3.Row], dict] = lambda row: dict(row)):

        res = FilterParser.parse_params(params, model_map)
        query, args = cls._apply_filter(query, **res)

        db = get_db()
        rows = db.execute(query, args).fetchall()
        df = pd.DataFrame([row_processor(row) for row in rows])

        if not df.empty:
            df.index = range(1, len(df) + 1)
            df.reset_index(inplace=True)
            df.columns = ['S/N'] + df.columns[1:].tolist()

        output_buffer = io.BytesIO()
        with pd.ExcelWriter(output_buffer) as writer:
            df.to_excel(writer, index=False)

        output_buffer.seek(0)
        wb = load_workbook(output_buffer)
        ws = wb.active

        for cell in ws[1]:
            cell.font = Font(bold=True)

        autofit_columns(ws, 55)
        new_buffer = io.BytesIO()
        wb.save(new_buffer)
        new_buffer.seek(0)
        return new_buffer

    @classmethod
    def download_facilities_sheet(cls, params: Params):
        query = f'''
        SELECT
            fc.name as 'Name',
            fc.local_government as 'Local Government',
            fc.ownership as Ownership,
            fc.facility_type as Type,
            GROUP_CONCAT(isc.scheme_name, ', ') as Scheme
        FROM facility as fc
        LEFT JOIN facility_scheme as fsc on fsc.facility_id = fc.id
        LEFT JOIN insurance_scheme as isc on fsc.scheme_id = isc.id
        '''
        model_map = {FacilityScheme: 'fsc', InsuranceScheme: 'isc', Facility: 'fc'}
        params = params.group(Facility, 'id')
        return cls.build_dataframe_buffer(query, params, model_map)

    @classmethod
    def download_encounter_sheet(cls, params: Params):
        query = f'''
        SELECT * FROM master_encounter_view
        '''
        return cls.build_dataframe_buffer(query, params, {})

    @classmethod
    def download_services_sheet(cls, params: Params):
        query = f'''
        SELECT
            srv.name as 'Name',
            sc.name as 'Category'
        FROM services as srv
        JOIN service_category as sc on sc.id = srv.category_id
        '''
        model_map = {Service: 'srv', ServiceCategory: 'sc'}
        return cls.build_dataframe_buffer(query, params, model_map)

    @classmethod
    def download_diseases_sheet(cls, params: Params):
        query = f'''
        SELECT
            dis.name as 'Name',
            cg.category_name as 'Category'
        FROM diseases as dis
        JOIN diseases_category as cg on cg.id = dis.category_id
        '''
        model_map = {Disease: 'dis', DiseaseCategory: 'cg'}
        return cls.build_dataframe_buffer(query, params, model_map)
